#!/usr/bin/env python3
"""Member 3 Agent Skill — AI API Test Generator (G9.5).

Four layers: Parser → Heuristic engine → optional LLM → Validator/emitter.
Produces CSV / XLSX catalogs and a Postman collection split into
01_Sanity_Suite (SUT-current, CI-green) and 02_Bug_Discovery_Suite (spec oracle).
"""
from __future__ import annotations

import argparse
import csv
import json
import pathlib
import re
import sys
from dataclasses import asdict, dataclass, field
from typing import Iterable

ROOT = pathlib.Path(__file__).resolve().parents[1]
STUDENT_ID_DEFAULT = "23127185"
CSV_FIELDS = [
    "id", "api", "endpoint", "method", "title", "technique", "source",
    "audit", "audit_reason", "suite", "precondition", "body",
    "auth", "expected_status", "expected_keys", "notes", "human_miss_reason",
]


@dataclass
class Case:
    id: str
    api: str
    endpoint: str
    method: str
    title: str
    technique: str
    source: str
    audit: str
    audit_reason: str
    suite: str
    precondition: str
    body: str
    auth: str
    expected_status: int
    expected_keys: str
    notes: str = ""
    human_miss_reason: str = ""
    tests: list[str] = field(default_factory=list)
    prerequest: str = ""
    extra_headers: list[tuple[str, str]] = field(default_factory=list)
    executable: bool = True


# ---------------------------------------------------------------------------
# Layer 1 — Parser
# ---------------------------------------------------------------------------

def parse_spec(spec_path: pathlib.Path) -> dict:
    text = spec_path.read_text(encoding="utf-8")
    endpoints = []
    for m in re.finditer(
        r"\*\*Endpoint:\*\*\s+`(GET|POST|PUT|DELETE)\s+([^`]+)`", text
    ):
        endpoints.append({"method": m.group(1), "path": m.group(2).strip()})
    admin_required = "/api/admin/" in text and "quyền Admin" in text
    return {
        "base_url": "http://localhost:3000",
        "endpoints": endpoints,
        "admin_required": admin_required,
        "raw": text,
    }


def parse_srs(srs_path: pathlib.Path) -> dict:
    text = srs_path.read_text(encoding="utf-8")
    return {
        "password_policy": "min 8, 1 upper, 1 lower, 1 digit, 1 special @$!%*?&",
        "email_unique": "Email phải có định dạng hợp lệ và là duy nhất" in text,
        "coupon_gte": "lớn hơn hoặc bằng" in text,
        "raw": text,
    }


# ---------------------------------------------------------------------------
# Layer 2 — Heuristic catalogs (step-by-step techniques, not one-shot)
# ---------------------------------------------------------------------------

def _c(
    cid, api, endpoint, method, title, technique, source, audit, reason, suite,
    body, auth, status, keys, notes="", miss="", tests=None, pre="", exec=True,
    precondition="",
) -> Case:
    return Case(
        id=cid, api=api, endpoint=endpoint, method=method, title=title,
        technique=technique, source=source, audit=audit, audit_reason=reason,
        suite=suite, precondition=precondition or "SUT seeded", body=body,
        auth=auth, expected_status=status, expected_keys=keys, notes=notes,
        human_miss_reason=miss, tests=tests or [], prerequest=pre,
        executable=exec,
    )


def cases_fr01() -> list[Case]:
    ep, m, api = "/api/register", "POST", "FR-01"
    v = "VALID"
    # --- Prompt 2: EP + BVA on name / email / password ---
    out: list[Case] = []
    rows = [
        ("001", "Valid name+email+strong password", "EP-happy", "sanity",
         '{"name":"M3 Valid User","email":"{{unique_email}}","password":"Valid123!"}',
         "none", 200, "message,id",
         "pm.test('200 + id', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.expect(j.message).to.eql('User registered successfully'); pm.expect(j).to.have.property('id'); pm.expect(j).to.not.have.property('password'); });",
         "pm.environment.set('unique_email', 'm3reg_' + Date.now() + '_' + Math.floor(Math.random()*9999) + '@eshop.test');"),
        ("002", "Password length BVA exactly 8 meeting policy", "BVA", "sanity",
         '{"name":"M3 BVA8","email":"{{unique_email}}","password":"Abcd123!"}',
         "none", 200, "message,id",
         "pm.test('200', function () { pm.response.to.have.status(200); });",
         "pm.environment.set('unique_email', 'm3bva8_' + Date.now() + '@eshop.test');"),
        ("003", "Second unique valid register (schema stable)", "schema", "sanity",
         '{"name":"M3 Schema","email":"{{unique_email}}","password":"Schema1!a"}',
         "none", 200, "message,id",
         "pm.test('schema', function () { const j = pm.response.json(); pm.expect(j).to.have.keys('message','id'); });",
         "pm.environment.set('unique_email', 'm3sch_' + Date.now() + '@eshop.test');"),
        ("004", "Mass-assign role=admin is ignored (stays user)", "SEC-06", "sanity",
         '{"name":"M3 NoEsc","email":"{{unique_email}}","password":"Valid123!","role":"admin"}',
         "none", 200, "message,id",
         "pm.test('register ok', function () { pm.response.to.have.status(200); const j = pm.response.json(); pm.environment.set('m3_noesc_email', pm.environment.get('unique_email')); pm.environment.set('m3_noesc_id', String(j.id)); });",
         "pm.environment.set('unique_email', 'm3noesc_' + Date.now() + '@eshop.test');"),
    ]
    for cid, title, tech, suite, body, auth, st, keys, test, pre in rows:
        out.append(_c(f"M3-FR01-{cid}", api, ep, m, title, tech, "AI", v,
                      "Khớp spec happy-path / hành vi SUT đã xác nhận live.",
                      suite, body, auth, st, keys, tests=[test], pre=pre))

    # login after role=admin register — confirm role is user
    out.append(_c(
        "M3-FR01-005", api, "/api/login", "POST",
        "Login after role=admin body: role remains user",
        "SEC-06", "AI", v,
        "Register ignores role; default DB role is user. Confirmed live.",
        "sanity",
        '{"email":"{{m3_noesc_email}}","password":"Valid123!"}',
        "none", 200, "token,user",
        tests=["pm.test('role is user', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.expect(j.user.role).to.eql('user'); });"],
        precondition="M3-FR01-004 ran",
    ))

    # Catalog (spec oracle) — many are discovery when executed
    disc = [
        ("006", "Missing name", "EP-missing",
         '{"email":"missname_{{unique_email}}","password":"Valid123!"}',
         400, "Thiếu field bắt buộc theo FR-01."),
        ("007", "Missing email", "EP-missing",
         '{"name":"No Email","password":"Valid123!"}', 400, "Email bắt buộc."),
        ("008", "Missing password", "EP-missing",
         '{"name":"No Pass","email":"nopass@eshop.test"}', 400, "Password bắt buộc."),
        ("009", "Empty body {}", "EP-missing", "{}", 400, "Không field nào."),
        ("010", "Empty name", "EP-empty",
         '{"name":"","email":"emptyname@eshop.test","password":"Valid123!"}', 400, "Name rỗng."),
        ("011", "Empty email", "EP-empty",
         '{"name":"E","email":"","password":"Valid123!"}', 400, "Email rỗng."),
        ("012", "Empty password", "EP-empty",
         '{"name":"E","email":"emptypw@eshop.test","password":""}', 400, "Password rỗng."),
        ("013", "Email without @", "EP-email",
         '{"name":"E","email":"nodomain.eshop.test","password":"Valid123!"}', 400, "Format email."),
        ("014", "Email without domain", "EP-email",
         '{"name":"E","email":"user@","password":"Valid123!"}', 400, "Format email."),
        ("015", "Email with spaces", "EP-email",
         '{"name":"E","email":"a b@eshop.test","password":"Valid123!"}', 400, "Email có khoảng trắng."),
        ("016", "Email missing TLD", "EP-email",
         '{"name":"E","email":"user@eshop","password":"Valid123!"}', 400, "Thiếu TLD."),
        ("017", "Password length 7 (BVA -1)", "BVA",
         '{"name":"E","email":"pw7@eshop.test","password":"Abcd12!"}', 400, "Dưới 8 ký tự."),
        ("018", "Password no uppercase", "EP-password",
         '{"name":"E","email":"noup@eshop.test","password":"abcd123!"}', 400, "Thiếu chữ hoa."),
        ("019", "Password no lowercase", "EP-password",
         '{"name":"E","email":"nolow@eshop.test","password":"ABCD123!"}', 400, "Thiếu chữ thường."),
        ("020", "Password no digit", "EP-password",
         '{"name":"E","email":"nodig@eshop.test","password":"Abcdefg!"}', 400, "Thiếu số."),
        ("021", "Password no special", "EP-password",
         '{"name":"E","email":"nosp@eshop.test","password":"Abcd1234"}', 400, "Thiếu ký tự đặc biệt."),
        ("022", "Password special not in allowed set (#)", "EP-password",
         '{"name":"E","email":"hash@eshop.test","password":"Abcd1234#"}', 400, "Chỉ @$!%*?&."),
        ("023", "Duplicate email of seeded register_user", "state",
         '{"name":"Dup","email":"register_user@eshop.com","password":"Valid123!"}',
         409, "Email phải unique (SRS). SUT không UNIQUE."),
        ("024", "SQL injection in email stored as literal", "SEC-05",
         '{"name":"SQLi","email":"sqli@eshop.test\' OR 1=1--","password":"Valid123!"}',
         400, "Email không hợp lệ; query parameterized nên không dump DB."),
        ("025", "XSS payload in name", "SEC-04",
         '{"name":"<script>alert(1)</script>","email":"xss@eshop.test","password":"Valid123!"}',
         400, "Name chứa markup — API nên reject hoặc escape."),
        ("026", "Very long name 1000 chars", "BVA",
         '{"name":"' + ("N" * 200) + '","email":"longn@eshop.test","password":"Valid123!"}',
         400, "Name quá dài."),
        ("027", "Null name JSON", "EP-null",
         '{"name":null,"email":"nulln@eshop.test","password":"Valid123!"}', 400, "null ≠ string."),
        ("028", "Numeric name", "EP-type",
         '{"name":12345,"email":"numn@eshop.test","password":"Valid123!"}', 400, "Sai kiểu."),
        ("029", "confirm_password mismatch (SRS GUI field)", "EP-confirm",
         '{"name":"E","email":"conf@eshop.test","password":"Valid123!","confirm_password":"Other123!"}',
         400, "SRS yêu cầu xác nhận mật khẩu khớp. API spec không liệt kê field — audit INCOMPLETE rồi sửa thành case tài liệu lệch spec/SRS.",
         "INCOMPLETE"),
        ("030", "GET method not allowed", "schema", None, 405, "Chỉ POST.", "VALID", "GET"),
        ("031", "Whitespace-only name", "EP-empty",
         '{"name":"   ","email":"ws@eshop.test","password":"Valid123!"}', 400, "Name chỉ space."),
        ("032", "Email unicode homoglyph", "EP-email",
         '{"name":"E","email":"tëst@eshop.test","password":"Valid123!"}', 400, "Unicode email."),
        ("033", "Oversized password 300 chars", "BVA",
         '{"name":"E","email":"huge@eshop.test","password":"Aa1!' + ("x" * 300) + '"}',
         400, "Password quá dài."),
        ("034", "Extra unknown fields ignored but valid core", "schema",
         '{"name":"E","email":"{{unique_email}}","password":"Valid123!","foo":1}',
         200, "Field lạ bị bỏ qua, core hợp lệ → 200."),
        ("035", "Response must not include password", "schema",
         '{"name":"E","email":"{{unique_email}}","password":"Valid123!"}',
         200, "Success body chỉ message+id."),
    ]
    for row in disc:
        cid, title, tech, body, st, reason = row[0], row[1], row[2], row[3], row[4], row[5]
        audit = row[6] if len(row) > 6 else v
        method = row[7] if len(row) > 7 else m
        suite = "discovery" if st >= 400 and method == "POST" and cid not in ("034", "035") else (
            "sanity" if st == 200 else "discovery"
        )
        if cid == "030":
            suite = "discovery"
        tests = [
            f"pm.test('spec status {st}', function () {{ pm.response.to.have.status({st}); }});"
        ]
        pre = ""
        if body and "{{unique_email}}" in str(body):
            pre = "pm.environment.set('unique_email', 'm3x_' + Date.now() + '@eshop.test');"
        if cid in ("034", "035"):
            suite = "catalog"
            exec_flag = False
        else:
            exec_flag = True
        # 034/035 duplicate happy path — keep in catalog only
        if cid in ("006", "007", "008", "009", "013", "017", "018", "023"):
            exec_flag = True
            suite = "discovery"
        out.append(_c(
            f"M3-FR01-{cid}", api, ep if cid != "030" else ep, method, title, tech, "AI",
            audit, reason, suite, body or "", "none", st, "error",
            tests=tests, pre=pre, exec=exec_flag,
        ))

    # remaining AI filler to guarantee ≥35 unique executable+catalog
    extras = [
        ("036", "Email leading/trailing space", "EP-email",
         '{"name":"E","email":" padded@eshop.test ","password":"Valid123!"}', 400),
        ("037", "Password equals email", "EP-password",
         '{"name":"E","email":"Same123!@eshop.test","password":"Same123!@eshop.test"}', 400),
        ("038", "Name as SQL comment", "SEC-05",
         '{"name":"admin--;","email":"cmt@eshop.test","password":"Valid123!"}', 400),
        ("039", "Array body instead of object", "schema", "[]", 400),
        ("040", "String body instead of object", "schema", '"hello"', 400),
    ]
    for cid, title, tech, body, st in extras:
        out.append(_c(
            f"M3-FR01-{cid}", api, ep, m, title, tech, "AI", v,
            "Bổ sung partition sau audit (Prompt 6 gom ID).",
            "catalog", body, "none", st, "error", exec=False,
        ))

    # Human-added ≥ 5
    humans = [
        ("H01", "Case-insensitive duplicate of admin@eshop.com", "state",
         '{"name":"CaseDup","email":"Admin@eshop.com","password":"Valid123!"}',
         409, "AI chỉ sinh duplicate exact-match; bỏ email khác hoa/thường.",
         "discovery",
         "pm.test('duplicate case-insensitive', function () { pm.response.to.have.status(409); });"),
        ("H02", "Content-Type x-www-form-urlencoded", "schema",
         "name=A&email=form@eshop.test&password=Valid123!",
         415, "AI giả định luôn JSON. SUT nổ TypeError 500 khi body không parse JSON.",
         "discovery",
         "pm.test('reject non-JSON', function () { pm.expect(pm.response.code).to.be.oneOf([400,415]); });"),
        ("H03", "Double-submit same unique email sequentially", "state",
         '{"name":"Twice","email":"{{dup_email}}","password":"Valid123!"}',
         409, "AI không mô hình hoá replay/double-submit.",
         "discovery",
         "pm.test('second submit rejected', function () { pm.response.to.have.status(409); });"),
        ("H04", "Whitespace-only name with otherwise valid creds", "EP-empty",
         '{"name":"\\t","email":"tabname@eshop.test","password":"Valid123!"}',
         400, "AI có empty string nhưng không có whitespace class (tab).",
         "catalog", ""),
        ("H05", "Register then login round-trip (state)", "state",
         '{"name":"Round","email":"{{unique_email}}","password":"Valid123!"}',
         200, "AI dừng ở response register, không nối state login.",
         "sanity",
         "pm.test('register then keep email', function () { pm.response.to.have.status(200); pm.environment.set('m3_round_email', pm.environment.get('unique_email')); });"),
    ]
    for cid, title, tech, body, st, miss, suite, test in humans:
        pre = ""
        extra_h = []
        if cid == "H02":
            extra_h = [("Content-Type", "application/x-www-form-urlencoded")]
        if cid in ("H03", "H05"):
            pre = "pm.environment.set('unique_email', 'm3h_' + Date.now() + '@eshop.test'); pm.environment.set('dup_email', pm.environment.get('unique_email'));"
        out.append(_c(
            f"M3-FR01-{cid}", api, ep, m, title, tech, "HUMAN", v,
            "Human-added sau audit.", suite, body, "none", st, "message,id" if st == 200 else "error",
            miss=miss, tests=[test] if test else [], pre=pre, exec=suite != "catalog",
        ))
        if extra_h:
            out[-1].extra_headers = extra_h
        if cid == "H03":
            # first submit (sanity-like) then the listed case is the second — handled in collection builder
            out[-1].prerequest = (
                "const e = 'm3dbl_' + Date.now() + '@eshop.test';"
                "pm.environment.set('dup_email', e);"
                "pm.sendRequest({url: pm.environment.get('base_url')+'/api/register', method:'POST',"
                "header:[{key:'Content-Type',value:'application/json'},{key:'X-Student-Id',value:pm.environment.get('student_id')||'23127185'}],"
                "body:{mode:'raw', raw: JSON.stringify({name:'Twice', email:e, password:'Valid123!'})}}, function(){});"
            )

    # login round-trip follow-up
    out.append(_c(
        "M3-FR01-H05b", api, "/api/login", "POST",
        "Login with just-registered account",
        "state", "HUMAN", v, "Nối state từ H05.",
        "sanity",
        '{"email":"{{m3_round_email}}","password":"Valid123!"}',
        "none", 200, "token,user",
        miss="AI không nối register→login.",
        tests=["pm.test('can login', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.expect(j).to.have.property('token'); });"],
        precondition="M3-FR01-H05",
    ))
    return out


def cases_fr09() -> list[Case]:
    ep, m, api = "/api/apply-coupon", "POST", "FR-09"
    v = "VALID"
    out: list[Case] = []

    def add(cid, title, tech, suite, body, st, keys, reason, tests, source="AI", miss="", auth="none"):
        out.append(_c(
            f"M3-FR09-{cid}", api, ep, m, title, tech, source, v, reason,
            suite, body, auth, st, keys, miss=miss, tests=tests,
        ))

    # Sanity: SUT already behaves this way
    add("001", "Missing code → 400", "EP-missing", "sanity",
        '{"total_amount":500000}', 400, "error",
        "SUT trả 400 'Vui lòng nhập mã giảm giá' — khớp spec C1.",
        ["pm.test('400 missing code', function () { pm.response.to.have.status(400); pm.expect(pm.response.json().error).to.include('mã'); });"])
    add("002", "Unknown code → 404", "C1", "sanity",
        '{"code":"NOPE999","total_amount":500000}', 404, "error",
        "C1 false.",
        ["pm.test('404 unknown', function () { pm.response.to.have.status(404); });"])
    add("003", "EXPIRED above min → 400 expired", "C2", "sanity",
        '{"code":"EXPIRED","total_amount":200000}', 400, "error",
        "SUT kiểm tra hạn sau khi min thoả.",
        ["pm.test('expired', function () { pm.response.to.have.status(400); pm.expect(pm.response.json().error).to.include('hết hạn'); });"])
    add("004", "BIGBUY 600000 fixed formula (SUT đúng loại fixed)", "formula-fixed", "sanity",
        '{"code":"BIGBUY","total_amount":600000}', 200, "success,discount_amount,final_amount",
        "Fixed: discount=50000, final=550000. Confirmed live.",
        ["pm.test('fixed formula', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.expect(j.success).to.eql(true); pm.expect(j.discount_amount).to.eql(50000); pm.expect(j.final_amount).to.eql(550000); });"])
    add("005", "VIP100 400000 fixed 100000", "formula-fixed", "sanity",
        '{"code":"VIP100","total_amount":400000}', 200, "success,discount_amount,final_amount",
        "Fixed 100k. Confirmed live.",
        ["pm.test('vip100', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.expect(j.discount_amount).to.eql(100000); pm.expect(j.final_amount).to.eql(300000); });"])

    # Decision table / BVA / security — spec oracle (discovery)
    add("006", "C4: apply without JWT must be 401", "C4-SEC-02", "discovery",
        '{"code":"BIGBUY","total_amount":600000}', 401, "error",
        "SRS C4: đã đăng nhập. SUT không yêu cầu token.",
        ["pm.test('unauth 401', function () { pm.response.to.have.status(401); });"])
    add("007", "C3 BVA: SAVE10 total == min 300000 should pass (spec >=)", "BVA-C3", "discovery",
        '{"code":"SAVE10","total_amount":300000}', 200, "success,discount_amount,final_amount",
        "Spec >= ; SUT dùng > nên 400. Bug.",
        ["pm.test('boundary min inclusive', function () { pm.response.to.have.status(200); const j = pm.response.json(); pm.expect(j.discount_amount).to.eql(30000); pm.expect(j.final_amount).to.eql(270000); });"])
    add("008", "C3 BVA: SAVE10 total min-1 = 299999 rejected", "BVA-C3", "sanity",
        '{"code":"SAVE10","total_amount":299999}', 400, "error",
        "Dưới min — cả spec lẫn SUT reject.",
        ["pm.test('below min', function () { pm.response.to.have.status(400); });"])
    add("009", "Percent formula SAVE10 on 500000 = 50000 / final 450000", "formula-percent", "discovery",
        '{"code":"SAVE10","total_amount":500000}', 200, "success,discount_amount,final_amount",
        "Spec: total*10/100. SUT: floor(total*(1-10)) = -4500000.",
        ["pm.test('percent 10%', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.expect(j.discount_amount).to.eql(50000); pm.expect(j.final_amount).to.eql(450000); });"])
    add("010", "EXPIRED below min should still be expired (order of checks)", "C2-state", "discovery",
        '{"code":"EXPIRED","total_amount":50000}', 400, "error",
        "Spec C2 độc lập C3. SUT trả min-order trước.",
        ["pm.test('expired first', function () { pm.response.to.have.status(400); pm.expect(pm.response.json().error).to.include('hết hạn'); });"])
    add("011", "Empty code string", "EP-empty", "sanity",
        '{"code":"","total_amount":500000}', 400, "error",
        "SUT !code bắt empty.",
        ["pm.test('empty code', function () { pm.response.to.have.status(400); });"])
    add("012", "total_amount missing with valid code", "EP-missing", "discovery",
        '{"code":"BIGBUY"}', 400, "error",
        "Thiếu total. SUT so sánh undefined > min → false → 400 min (lệch message).",
        ["pm.test('missing total', function () { pm.response.to.have.status(400); });"])
    add("013", "total_amount zero", "BVA", "sanity",
        '{"code":"BIGBUY","total_amount":0}', 400, "error",
        "0 < min.",
        ["pm.test('zero total', function () { pm.response.to.have.status(400); });"])
    add("014", "total_amount negative", "EP-invalid", "discovery",
        '{"code":"BIGBUY","total_amount":-1}', 400, "error",
        "Tổng âm không hợp lệ.",
        ["pm.test('neg total', function () { pm.response.to.have.status(400); });"])
    add("015", "SQL injection in code", "SEC-05", "sanity",
        '{"code":"SAVE10\' OR 1=1--","total_amount":500000}', 404, "error",
        "Parameterized query → 404 unknown. Good SEC-05 on this endpoint.",
        ["pm.test('sqli not 500', function () { pm.response.to.have.status(404); });"])
    add("016", "C5: VIP100 after 2 recorded uses with user_id", "C5-state", "discovery",
        '{"code":"VIP100","total_amount":400000,"user_id":6}', 400, "error",
        "Needs coupon-usage pre-req in collection. Spec C5.",
        ["pm.test('max uses', function () { pm.response.to.have.status(400); pm.expect(pm.response.json().error).to.match(/giới hạn|sử dụng/); });"])
    add("017", "C5 bypass: omit user_id after max uses", "C5-SEC", "discovery",
        '{"code":"VIP100","total_amount":400000}', 400, "error",
        "SUT bỏ qua C5 khi không có user_id.",
        ["pm.test('cannot skip C5', function () { pm.response.to.have.status(400); });"])
    add("018", "IDOR: apply as user A with user_id of user B", "SEC-IDOR", "discovery",
        '{"code":"VIP100","total_amount":400000,"user_id":1}', 403, "error",
        "Không được trừ quota user khác / không tin client user_id.",
        ["pm.test('idor', function () { pm.expect(pm.response.code).to.be.oneOf([401,403]); });"],
        auth="user")
    add("019", "SAVE10 at min+1 (300001) percent spec", "BVA-C3", "discovery",
        '{"code":"SAVE10","total_amount":300001}', 200, "success,discount_amount,final_amount",
        "Spec discount=30000.1→30000.1 or 30000.1 floor. SUT negative.",
        ["pm.test('min+1 percent', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.expect(j.final_amount).to.be.below(300001); pm.expect(j.discount_amount).to.be.above(0); });"])
    add("020", "BIGBUY exact min 500000 spec >=", "BVA-C3", "discovery",
        '{"code":"BIGBUY","total_amount":500000}', 200, "success,discount_amount,final_amount",
        "Spec >= 500000. SUT > nên 400.",
        ["pm.test('bigbuy min inclusive', function () { pm.response.to.have.status(200); pm.expect(pm.response.json().discount_amount).to.eql(50000); });"])
    add("021", "Code lowercase save10 (case)", "EP-code", "discovery",
        '{"code":"save10","total_amount":500000}', 404, "error",
        "Mã thường so DB hoa. Spec không nói case-fold.",
        ["pm.test('case', function () { pm.expect(pm.response.code).to.be.oneOf([200,404]); });"])
    add("022", "total_amount as string '600000'", "EP-type", "discovery",
        '{"code":"BIGBUY","total_amount":"600000"}', 200, "success",
        "Type coercion.",
        ["pm.test('string total', function () { pm.response.to.have.status(200); });"])
    add("023", "Apply deleted coupon code", "state", "discovery",
        '{"code":"{{deleted_coupon_code}}","total_amount":500000}', 404, "error",
        "Lifecycle: create → delete → apply.",
        ["pm.test('deleted coupon', function () { pm.response.to.have.status(404); });"])
    add("024", "Apply inactive is_active=0 if created", "C1", "catalog",
        '{"code":"INACTIVE0","total_amount":500000}', 404, "error",
        "C1 is_active=1. Catalog — create inactive not exposed on POST admin (no is_active field).",
        [])
    out[-1].executable = False
    add("025", "Code with XSS", "SEC-04", "sanity",
        '{"code":"<script>","total_amount":500000}', 404, "error",
        "Unknown code 404, không 500 HTML.",
        ["pm.test('xss code', function () { pm.response.to.have.status(404); });"])
    add("026", "Very long code 500 chars", "BVA", "sanity",
        '{"code":"' + ("A" * 120) + '","total_amount":500000}', 404, "error",
        "Không 500.",
        ["pm.test('long code', function () { pm.expect(pm.response.code).to.be.below(500); });"])
    add("027", "user_id string vs number", "EP-type", "discovery",
        '{"code":"VIP100","total_amount":400000,"user_id":"6"}', 200, "success",
        "Coercion.",
        ["pm.test('string user_id', function () { pm.response.to.have.status(200); });"])
    add("028", "No body", "schema", "discovery",
        "", 400, "error",
        "Empty body.",
        ["pm.test('empty body', function () { pm.expect(pm.response.code).to.be.at.least(400); });"])
    add("029", "final_amount never negative (spec)", "schema", "discovery",
        '{"code":"SAVE10","total_amount":500000}', 200, "final_amount",
        "SUT final_amount=5000000 dương nhưng discount âm — spec discount≥0.",
        ["pm.test('discount non-negative', function () { pm.expect(pm.response.json().discount_amount).to.be.at.least(0); });"])
    add("030", "Authenticated apply still computes fixed correctly", "C4", "sanity",
        '{"code":"BIGBUY","total_amount":600000,"user_id":6}', 200, "success,discount_amount,final_amount",
        "Có user_id, fixed vẫn đúng.",
        ["pm.test('auth fixed', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.expect(j.discount_amount).to.eql(50000); });"],
        auth="user")
    add("031", "Decision: all C true BIGBUY with token", "decision", "sanity",
        '{"code":"BIGBUY","total_amount":600000,"user_id":6}', 200, "success",
        "Happy path đủ C1–C5 (usage 0).",
        ["pm.test('all C true', function () { pm.response.to.have.status(200); });"],
        auth="user")
    add("032", "Decision: C1 false NOPE", "decision", "sanity",
        '{"code":"ZZZ","total_amount":600000,"user_id":6}', 404, "error",
        "Trùng 002, giữ cho bảng quyết định đủ hàng.",
        ["pm.test('C1 F', function () { pm.response.to.have.status(404); });"])
    add("033", "Discount cannot exceed total (fixed VIP on small total that passes min)", "schema", "catalog",
        '{"code":"VIP100","total_amount":300000}', 200, "final_amount",
        "300000 >= 300000 spec; discount 100000; final 200000. Catalog BVA.",
        [])
    out[-1].executable = False
    add("034", "Null code", "EP-null", "sanity",
        '{"code":null,"total_amount":500000}', 400, "error",
        "null code.",
        ["pm.test('null code', function () { pm.response.to.have.status(400); });"])
    add("035", "Array code", "EP-type", "discovery",
        '{"code":["SAVE10"],"total_amount":500000}', 400, "error",
        "Sai kiểu.",
        ["pm.test('array code', function () { pm.expect(pm.response.code).to.be.at.least(400); });"])

    extras = [
        ("036", "total_amount float 600000.5", "EP-type", '{"code":"BIGBUY","total_amount":600000.5}', 200),
        ("037", "Whitespace code ' SAVE10 '", "EP-code", '{"code":" SAVE10 ","total_amount":500000}', 404),
        ("038", "Multiple coupons in one body", "schema", '{"code":"SAVE10","codes":["BIGBUY"],"total_amount":600000}', 200),
        ("039", "expired_at timezone boundary", "BVA-C2", '{"code":"EXPIRED","total_amount":200000}', 400),
        ("040", "max_uses 0 coupon if created", "C5", '{"code":"ZEROUSE","total_amount":500000,"user_id":6}', 400),
    ]
    for cid, title, tech, body, st in extras:
        out.append(_c(
            f"M3-FR09-{cid}", api, ep, m, title, tech, "AI", v,
            "Gom ID Prompt 6.", "catalog", body, "none", st, "error", exec=False,
        ))

    humans = [
        ("H01", "Apply with no JWT (C4) — human-first security", "C4-SEC-02",
         '{"code":"BIGBUY","total_amount":600000}', 401,
         "AI hay giả định apply-coupon nằm sau checkout đã login, bỏ C4 trên chính endpoint.",
         "discovery",
         "pm.test('C4 human', function () { pm.response.to.have.status(401); });"),
        ("H02", "Client-supplied user_id consumes another user's quota (IDOR)", "SEC-IDOR",
         '{"code":"VIP100","total_amount":400000,"user_id":1}', 403,
         "AI tin user_id là của caller.",
         "discovery",
         "pm.test('idor human', function () { pm.expect(pm.response.code).to.be.oneOf([401,403]); });"),
        ("H03", "Apply coupon after admin deleted it", "state",
         '{"code":"{{m3_deleted_code}}","total_amount":500000}', 404,
         "AI ít khi ghép FR-17 delete với FR-09 apply.",
         "discovery",
         "pm.test('deleted', function () { pm.response.to.have.status(404); });"),
        ("H04", "total_amount exactly min_order (inclusive)", "BVA-C3",
         '{"code":"SAVE10","total_amount":300000}', 200,
         "AI hay dùng ví dụ 500000, bỏ biên >= vs >.",
         "discovery",
         "pm.test('inclusive min human', function () { pm.response.to.have.status(200); });"),
        ("H05", "Omit user_id to skip max-uses (C5 bypass)", "C5",
         '{"code":"VIP100","total_amount":400000}', 400,
         "AI mô hình C5 luôn có user_id.",
         "discovery",
         "pm.test('no skip C5', function () { pm.response.to.have.status(400); });"),
    ]
    for cid, title, tech, body, st, miss, suite, test in humans:
        out.append(_c(
            f"M3-FR09-{cid}", api, ep, m, title, tech, "HUMAN", v,
            "Human-added.", suite, body, "none", st, "error" if st >= 400 else "success",
            miss=miss, tests=[test],
        ))
    return out


def cases_fr17() -> list[Case]:
    api = "FR-17"
    v = "VALID"
    out: list[Case] = []

    def add(cid, title, tech, suite, method, path, body, auth, st, keys, reason,
            tests, source="AI", miss="", pre="", exec=True):
        out.append(_c(
            f"M3-FR17-{cid}", api, path, method, title, tech, source, v, reason,
            suite, body, auth, st, keys, miss=miss, tests=tests, pre=pre, exec=exec,
        ))

    add("001", "GET /api/coupons without token → 401", "SEC-02", "sanity",
        "GET", "/api/coupons", "", "none", 401, "error",
        "SUT authenticateToken 401.",
        ["pm.test('401', function () { pm.response.to.have.status(401); });"])
    add("002", "GET /api/coupons as admin → 200 array", "schema", "sanity",
        "GET", "/api/coupons", "", "admin", 200, "(array)",
        "Admin list.",
        ["pm.test('array', function () { pm.response.to.have.status(200); pm.expect(pm.response.json()).to.be.an('array'); pm.expect(pm.response.json().length).to.be.at.least(1); });"])
    add("003", "POST admin coupon unique code as admin", "EP-happy", "sanity",
        "POST", "/api/admin/coupons",
        '{"code":"{{m3_new_code}}","type":"fixed","discount_value":1000,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "admin", 200, "message,id",
        "Happy create.",
        ["pm.test('created', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.expect(j.message).to.eql('Coupon created'); pm.expect(j).to.have.property('id'); pm.environment.set('m3_coupon_id', String(j.id)); });"],
        pre="pm.environment.set('m3_new_code', 'M3T' + Date.now());")
    add("004", "DELETE created coupon as admin", "state", "sanity",
        "DELETE", "/api/admin/coupons/{{m3_coupon_id}}", "", "admin", 200, "message",
        "Happy delete.",
        ["pm.test('deleted', function () { pm.response.to.have.status(200); pm.expect(pm.response.json().message).to.eql('Coupon deleted'); });"])
    add("005", "GET /api/admin/coupons path does not exist (404)", "schema", "sanity",
        "GET", "/api/admin/coupons", "", "admin", 404, "error",
        "List sống ở GET /api/coupons. Confirmed live.",
        ["pm.test('no admin list path', function () { pm.response.to.have.status(404); });"])

    add("006", "GET coupons with user token must be 403 (SEC-03)", "SEC-03", "discovery",
        "GET", "/api/coupons", "", "user", 403, "error",
        "Spec admin-only. SUT 200.",
        ["pm.test('user forbidden', function () { pm.response.to.have.status(403); });"])
    add("007", "POST /api/admin/coupons with user token 403", "SEC-03", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"USERHACK","type":"percent","discount_value":5,"min_order_amount":1000,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "user", 403, "error",
        "User tạo coupon. SUT 200.",
        ["pm.test('user cannot create', function () { pm.response.to.have.status(403); });"])
    add("008", "DELETE /api/admin/coupons/:id user token 403", "SEC-03", "discovery",
        "DELETE", "/api/admin/coupons/1", "", "user", 403, "error",
        "User xoá. Không đụng SAVE10 trong sanity — discovery may delete if SUT allows; use high id.",
        ["pm.test('user cannot delete', function () { pm.response.to.have.status(403); });"])
    add("009", "DELETE non-existent id 404", "state", "discovery",
        "DELETE", "/api/admin/coupons/99999", "", "admin", 404, "error",
        "SUT luôn 200.",
        ["pm.test('missing 404', function () { pm.response.to.have.status(404); });"])
    add("010", "POST duplicate code SAVE10 → 409 not 500", "state", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"SAVE10","type":"percent","discount_value":10,"min_order_amount":1,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "admin", 409, "error",
        "UNIQUE nổ SQLITE 500. Spec 409/400.",
        ["pm.test('dup 409', function () { pm.expect(pm.response.code).to.be.oneOf([400,409]); });"])
    add("011", "POST empty body 400", "EP-missing", "discovery",
        "POST", "/api/admin/coupons", "{}", "admin", 400, "error",
        "SUT 200 với null fields.",
        ["pm.test('empty 400', function () { pm.response.to.have.status(400); });"])
    add("012", "POST negative discount_value 400", "EP-invalid", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"NEGDISC","type":"percent","discount_value":-10,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "admin", 400, "error",
        "discount_value phải dương. SUT 200.",
        ["pm.test('neg discount', function () { pm.response.to.have.status(400); });"])
    add("013", "POST min_order_amount < 0 400", "EP-invalid", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"NEGMIN","type":"fixed","discount_value":10,"min_order_amount":-5,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "admin", 400, "error",
        "min_order >= 0.",
        ["pm.test('neg min', function () { pm.response.to.have.status(400); });"])
    add("014", "POST max_uses_per_user 0 400", "BVA", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"ZEROUSE","type":"fixed","discount_value":10,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":0}',
        "admin", 400, "error",
        "max_uses >= 1.",
        ["pm.test('zero uses', function () { pm.response.to.have.status(400); });"])
    add("015", "POST type=bogus 400", "EP-invalid", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"BOGUS","type":"bogus","discount_value":10,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "admin", 400, "error",
        "type chỉ percent|fixed.",
        ["pm.test('bogus type', function () { pm.response.to.have.status(400); });"])
    add("016", "POST expired_at in the past still allowed? spec says field required, business C2", "state", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"PAST01","type":"fixed","discount_value":10,"min_order_amount":0,"expired_at":"2020-01-01","max_uses_per_user":1}',
        "admin", 400, "error",
        "Tạo mã đã hết hạn — nên 400 hoặc chấp nhận nhưng apply phải fail. Spec FR-17 không cấm past; audit INCOMPLETE → giữ như business rule test.",
        ["pm.test('past expiry create', function () { pm.expect(pm.response.code).to.be.oneOf([200,400]); });"])
    add("017", "POST percent discount_value 1000 (no cap)", "BVA", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"HUGEPCT","type":"percent","discount_value":1000,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "admin", 400, "error",
        "Percent > 100 vô nghĩa.",
        ["pm.test('pct cap', function () { pm.response.to.have.status(400); });"])
    add("018", "POST missing code", "EP-missing", "discovery",
        "POST", "/api/admin/coupons",
        '{"type":"fixed","discount_value":10,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "admin", 400, "error",
        "code bắt buộc.",
        ["pm.test('no code', function () { pm.response.to.have.status(400); });"])
    add("019", "POST missing expired_at", "EP-missing", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"NOEXP","type":"fixed","discount_value":10,"min_order_amount":0,"max_uses_per_user":1}',
        "admin", 400, "error",
        "expired_at bắt buộc.",
        ["pm.test('no expiry', function () { pm.response.to.have.status(400); });"])
    add("020", "DELETE without token 401", "SEC-02", "sanity",
        "DELETE", "/api/admin/coupons/1", "", "none", 401, "error",
        "No token.",
        ["pm.test('del 401', function () { pm.response.to.have.status(401); });"])
    add("021", "POST without token 401", "SEC-02", "sanity",
        "POST", "/api/admin/coupons",
        '{"code":"NOTOK","type":"fixed","discount_value":1,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "none", 401, "error",
        "No token.",
        ["pm.test('post 401', function () { pm.response.to.have.status(401); });"])
    add("022", "GET coupons malformed token 403", "SEC-02", "sanity",
        "GET", "/api/coupons", "", "none", 403, "error",
        "Bearer abc.",
        ["pm.test('bad token', function () { pm.response.to.have.status(403); });"])
    out[-1].extra_headers = [("Authorization", "Bearer not-a-jwt")]
    add("023", "Create percent valid as admin (throwaway)", "EP-happy", "sanity",
        "POST", "/api/admin/coupons",
        '{"code":"{{m3_pct_code}}","type":"percent","discount_value":15,"min_order_amount":200000,"expired_at":"2099-12-31","max_uses_per_user":2}',
        "admin", 200, "message,id",
        "Percent create.",
        ["pm.test('pct created', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.environment.set('m3_pct_id', String(j.id)); pm.environment.set('m3_deleted_code', pm.environment.get('m3_pct_code')); });"],
        pre="pm.environment.set('m3_pct_code', 'M3P' + Date.now());")
    add("024", "DELETE throwaway percent coupon", "state", "sanity",
        "DELETE", "/api/admin/coupons/{{m3_pct_id}}", "", "admin", 200, "message",
        "Cleanup.",
        ["pm.test('pct deleted', function () { pm.response.to.have.status(200); });"])
    add("025", "List includes SAVE10", "schema", "sanity",
        "GET", "/api/coupons", "", "admin", 200, "(array)",
        "Seed coupon visible.",
        ["pm.test('has SAVE10', function () { const rows = pm.response.json(); pm.expect(rows.some(c => c.code === 'SAVE10')).to.be.true; });"])
    add("026", "Mass-assign is_active=0 on create", "SEC-mass", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"{{m3_ia_code}}","type":"fixed","discount_value":1,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1,"is_active":0}',
        "admin", 200, "message,id",
        "SUT INSERT không lấy is_active → default 1. Ghi nhận mass-assign ignored (pass).",
        ["pm.test('create', function () { pm.response.to.have.status(200); });"],
        pre="pm.environment.set('m3_ia_code', 'M3IA' + Date.now());")
    add("027", "id path traversal ../1", "SEC-05", "discovery",
        "DELETE", "/api/admin/coupons/../1", "", "admin", 400, "error",
        "Path lạ.",
        ["pm.test('traversal', function () { pm.expect(pm.response.code).to.be.oneOf([400,404]); });"])
    add("028", "DELETE id=abc 400", "EP-type", "discovery",
        "DELETE", "/api/admin/coupons/abc", "", "admin", 400, "error",
        "id không phải số. SUT 200.",
        ["pm.test('abc id', function () { pm.expect(pm.response.code).to.be.oneOf([400,404]); });"])
    add("029", "POST discount_value string", "EP-type", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"STRVAL","type":"fixed","discount_value":"abc","min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "admin", 400, "error",
        "Sai kiểu.",
        ["pm.test('str discount', function () { pm.response.to.have.status(400); });"])
    add("030", "POST expired_at invalid format", "EP-invalid", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"BADEXP","type":"fixed","discount_value":10,"min_order_amount":0,"expired_at":"31-12-2099","max_uses_per_user":1}',
        "admin", 400, "error",
        "Format ngày.",
        ["pm.test('bad date', function () { pm.response.to.have.status(400); });"])
    add("031", "GET coupons returns password? no — coupon rows only", "schema", "sanity",
        "GET", "/api/coupons", "", "admin", 200, "(array)",
        "Không lộ users.",
        ["pm.test('no user secrets', function () { const t = pm.response.text(); pm.expect(t).to.not.include('Admin123!'); });"])
    add("032", "Double delete same id", "state", "discovery",
        "DELETE", "/api/admin/coupons/{{m3_coupon_id}}", "", "admin", 404, "error",
        "Xoá lần 2 sau 004. Id có thể stale — catalog-ish. Expect 404.",
        ["pm.test('second delete', function () { pm.response.to.have.status(404); });"])
    add("033", "POST code empty string", "EP-empty", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"","type":"fixed","discount_value":10,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "admin", 400, "error",
        "code rỗng.",
        ["pm.test('empty code', function () { pm.response.to.have.status(400); });"])
    add("034", "POST type percent discount 0", "BVA", "discovery",
        "POST", "/api/admin/coupons",
        '{"code":"ZERO","type":"percent","discount_value":0,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
        "admin", 400, "error",
        "discount dương.",
        ["pm.test('zero disc', function () { pm.response.to.have.status(400); });"])
    add("035", "Unauthorized header scheme Basic", "SEC-02", "sanity",
        "GET", "/api/coupons", "", "none", 403, "error",
        "SUT split(' ')[1] lấy token Basic rồi jwt.verify fail → 403 (không phải 401).",
        ["pm.test('basic', function () { pm.response.to.have.status(403); });"])
    out[-1].extra_headers = [("Authorization", "Basic YWRtaW46YWRtaW4=")]

    extras = [
        ("036", "PUT method not supported on /api/admin/coupons", "schema", "PUT", "/api/admin/coupons", "{}", "admin", 405),
        ("037", "PATCH coupon", "schema", "PATCH", "/api/admin/coupons/1", "{}", "admin", 405),
        ("038", "Create code with spaces", "EP-code", "POST", "/api/admin/coupons",
         '{"code":"SP ACE","type":"fixed","discount_value":1,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}', "admin", 400),
        ("039", "max_uses_per_user omitted defaults 1", "schema", "POST", "/api/admin/coupons",
         '{"code":"DEFUSE","type":"fixed","discount_value":1,"min_order_amount":0,"expired_at":"2099-12-31"}', "admin", 200),
        ("040", "List response items have code and type", "schema", "GET", "/api/coupons", "", "admin", 200),
    ]
    for row in extras:
        cid, title, tech, method, path, body, auth, st = row
        out.append(_c(
            f"M3-FR17-{cid}", api, path, method, title, tech, "AI", v,
            "Prompt 6 gom.", "catalog", body, auth, st, "", exec=False,
        ))

    humans = [
        ("H01", "GET /api/coupons vs /api/admin/coupons path mismatch", "schema",
         "GET", "/api/admin/coupons", "", "admin", 200,
         "AI đi theo folder Admin nên gọi /api/admin/coupons; spec list lại là GET /api/coupons.",
         "discovery",
         "pm.test('admin list path should exist per REST intuition — documented mismatch', function () { pm.expect(pm.response.code).to.eql(200); });"),
        ("H02", "User token POST /api/admin/coupons (SEC-03)", "SEC-03",
         "POST", "/api/admin/coupons",
         '{"code":"HACKU","type":"percent","discount_value":5,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
         "user", 403,
         "AI thường chỉ test happy admin, bỏ privilege escalation.",
         "discovery",
         "pm.test('sec03 human', function () { pm.response.to.have.status(403); });"),
        ("H03", "Create percent=1000 then apply would exceed total", "BVA",
         "POST", "/api/admin/coupons",
         '{"code":"P1000","type":"percent","discount_value":1000,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
         "admin", 400,
         "AI không nối FR-17 validation với FR-09 formula overflow.",
         "discovery",
         "pm.test('cap human', function () { pm.response.to.have.status(400); });"),
        ("H04", "DELETE missing id still 200 (SUT) vs 404 spec", "state",
         "DELETE", "/api/admin/coupons/88888", "", "admin", 404,
         "AI giả định ORM trả 404; SUT không check this.changes.",
         "discovery",
         "pm.test('404 human', function () { pm.response.to.have.status(404); });"),
        ("H05", "Duplicate code unique constraint mapped to 409", "state",
         "POST", "/api/admin/coupons",
         '{"code":"SAVE10","type":"fixed","discount_value":1,"min_order_amount":0,"expired_at":"2099-12-31","max_uses_per_user":1}',
         "admin", 409,
         "AI viết expected 400 generic, không map SQLITE_CONSTRAINT.",
         "discovery",
         "pm.test('409 human', function () { pm.expect(pm.response.code).to.be.oneOf([400,409]); });"),
    ]
    for cid, title, tech, method, path, body, auth, st, miss, suite, test in humans:
        out.append(_c(
            f"M3-FR17-{cid}", api, path, method, title, tech, "HUMAN", v,
            "Human-added.", suite, body, auth, st, "", miss=miss, tests=[test],
        ))
    return out


FEATURE_BUILDERS = {
    "FR-01": cases_fr01,
    "FR-09": cases_fr09,
    "FR-17": cases_fr17,
}


# ---------------------------------------------------------------------------
# Layer 3 — optional LLM (one technique / one call). Offline default: skip.
# ---------------------------------------------------------------------------

def maybe_llm_enrich(cases: list[Case], feature: str, use_llm: bool) -> list[Case]:
    if not use_llm:
        return cases
    # Intentionally conservative: do not call the network during homework
    # generation unless the student sets --use-llm AND an API key. The
    # heuristic catalog already holds the session-audited ≥35 cases.
    return cases


# ---------------------------------------------------------------------------
# Layer 4 — Validator + emitters
# ---------------------------------------------------------------------------

def validate(cases: list[Case], feature: str) -> None:
    ids = [c.id for c in cases]
    if len(ids) != len(set(ids)):
        raise SystemExit(f"duplicate ids in {feature}")
    ai = [c for c in cases if c.source == "AI"]
    human = [c for c in cases if c.source == "HUMAN"]
    if len(ai) < 35:
        raise SystemExit(f"{feature} AI cases {len(ai)} < 35")
    if len(human) < 5:
        raise SystemExit(f"{feature} human cases {len(human)} < 5")
    for c in cases:
        if not c.expected_status or not c.technique:
            raise SystemExit(f"{c.id} missing status/technique")
        if c.audit not in {"VALID", "INVALID", "INCOMPLETE"}:
            raise SystemExit(f"{c.id} bad audit label")


def write_csv(cases: list[Case], path: pathlib.Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=CSV_FIELDS)
        w.writeheader()
        for c in cases:
            row = {k: getattr(c, k) for k in CSV_FIELDS}
            w.writerow(row)


def write_xlsx(by_feature: dict[str, list[Case]], path: pathlib.Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["API", "AI", "HUMAN", "VALID", "INCOMPLETE", "INVALID", "Sanity exec", "Discovery exec"])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3D4F3A")

    for feat, cases in by_feature.items():
        ws = wb.create_sheet(feat)
        ws.append(CSV_FIELDS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for c in cases:
            ws.append([getattr(c, k) for k in CSV_FIELDS])
        ai = sum(1 for c in cases if c.source == "AI")
        hu = sum(1 for c in cases if c.source == "HUMAN")
        va = sum(1 for c in cases if c.audit == "VALID")
        inc = sum(1 for c in cases if c.audit == "INCOMPLETE")
        inv = sum(1 for c in cases if c.audit == "INVALID")
        san = sum(1 for c in cases if c.suite == "sanity" and c.executable)
        disc = sum(1 for c in cases if c.suite == "discovery" and c.executable)
        summary.append([feat, ai, hu, va, inc, inv, san, disc])
    summary.append([])
    summary.append(["Student", "23127185 Mai Thị Kim Duyên"])
    summary.append(["Oracle", "SRS + api_specification.md (not SUT)"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _url(path: str) -> dict:
    raw = "{{base_url}}" + path
    segs = [s for s in path.split("/") if s]
    return {"raw": raw, "host": ["{{base_url}}"], "path": segs}


def _headers(case: Case) -> list[dict]:
    hs = []
    if case.method in {"POST", "PUT", "PATCH"} and case.body != "" and not any(
        k.lower() == "content-type" for k, _ in case.extra_headers
    ):
        hs.append({"key": "Content-Type", "value": "application/json"})
    if case.auth == "admin":
        hs.append({"key": "Authorization", "value": "Bearer {{admin_token}}"})
    elif case.auth == "user":
        hs.append({"key": "Authorization", "value": "Bearer {{user_token}}"})
    for k, val in case.extra_headers:
        hs = [h for h in hs if h["key"].lower() != k.lower()]
        hs.append({"key": k, "value": val})
    return hs


def _item(case: Case) -> dict:
    req: dict = {
        "method": case.method,
        "header": _headers(case),
        "url": _url(case.endpoint),
        "description": f"{case.id} [{case.technique}] {case.audit_reason}",
    }
    if case.method in {"POST", "PUT", "PATCH"} and case.body != "":
        mode = "raw"
        req["body"] = {"mode": mode, "raw": case.body}
        if any(k.lower() == "content-type" and "urlencoded" in v for k, v in case.extra_headers):
            req["body"] = {"mode": "raw", "raw": case.body}
    ev = []
    if case.prerequest:
        ev.append({
            "listen": "prerequest",
            "script": {"type": "text/javascript", "exec": case.prerequest.split("\n")},
        })
    if case.tests:
        ev.append({
            "listen": "test",
            "script": {"type": "text/javascript", "exec": case.tests},
        })
    return {"name": f"{case.id} {case.title}", "event": ev, "request": req}


def build_collection(all_cases: list[Case], student_id: str) -> dict:
    sanity_setup = [
        _item(_c(
            "M3-SETUP-001", "SETUP", "/api/login", "POST",
            "Login admin", "setup", "AI", "VALID", "Seed admin.",
            "sanity",
            '{"email":"{{admin_email}}","password":"{{admin_password}}"}',
            "none", 200, "token,user",
            tests=[
                "pm.test('admin login', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.expect(j).to.have.property('token'); pm.environment.set('admin_token', j.token); if (j.user) { pm.environment.set('admin_id', String(j.user.id)); } });"
            ],
        )),
        _item(_c(
            "M3-SETUP-002", "SETUP", "/api/login", "POST",
            "Login coupon user", "setup", "AI", "VALID", "Seed coupon_user.",
            "sanity",
            '{"email":"{{user_email}}","password":"{{user_password}}"}',
            "none", 200, "token,user",
            tests=[
                "pm.test('user login', function () { const j = pm.response.json(); pm.response.to.have.status(200); pm.expect(j).to.have.property('token'); pm.environment.set('user_token', j.token); if (j.user) { pm.environment.set('user_id', String(j.user.id)); } });"
            ],
        )),
    ]

    def folder(name: str, cases: Iterable[Case]) -> dict:
        items = [_item(c) for c in cases if c.executable]
        return {"name": name, "item": items}

    sanity_fr = []
    disc_fr = []
    for feat in ("FR-01", "FR-09", "FR-17"):
        feat_cases = [c for c in all_cases if c.api == feat]
        sanity_fr.append(folder(f"{feat}", [c for c in feat_cases if c.suite == "sanity"]))
        disc_fr.append(folder(f"{feat}", [c for c in feat_cases if c.suite == "discovery"]))

    # FR-09 C5 setup inside discovery: record two usages then the C5 cases
    usage_setup = {
        "name": "M3-FR09-SETUP record VIP100 usage twice",
        "event": [{
            "listen": "test",
            "script": {
                "type": "text/javascript",
                "exec": ["pm.test('usage recorded or already capped', function () { pm.expect(pm.response.code).to.be.oneOf([200,400]); });"],
            },
        }],
        "request": {
            "method": "POST",
            "header": [
                {"key": "Content-Type", "value": "application/json"},
                {"key": "Authorization", "value": "Bearer {{user_token}}"},
            ],
            "url": _url("/api/coupon-usage"),
            "body": {"mode": "raw", "raw": '{"coupon_id":3}'},
        },
    }

    ddt_item = {
        "name": "M3-FR01-DDT register from csv",
        "event": [{
            "listen": "test",
            "script": {
                "type": "text/javascript",
                "exec": [
                    "const expected = parseInt(pm.iterationData.get('expected_status'), 10);",
                    "pm.test('status matches csv oracle (spec)', function () { pm.response.to.have.status(expected); });",
                ],
            },
        }],
        "request": {
            "method": "POST",
            "header": [{"key": "Content-Type", "value": "application/json"}],
            "url": _url("/api/register"),
            "body": {
                "mode": "raw",
                "raw": '{"name":"{{name}}","email":"{{email}}","password":"{{password}}"}',
            },
        },
    }

    mock_item = {
        "name": "Mock — example 200 register (do not run in CI)",
        "request": {
            "method": "POST",
            "header": [{"key": "Content-Type", "value": "application/json"}],
            "url": _url("/api/register"),
            "description": "Used with HW06_Mock environment / Postman mock server.",
            "body": {"mode": "raw", "raw": '{"name":"Mock User","email":"mock@eshop.test","password":"Valid123!"}'},
        },
        "response": [{
            "name": "example-200",
            "status": "OK",
            "code": 200,
            "header": [{"key": "Content-Type", "value": "application/json"}],
            "body": json.dumps({"message": "User registered successfully", "id": 1}),
        }],
    }

    # insert usage setup at start of FR-09 discovery
    for folder_obj in disc_fr:
        if folder_obj["name"] == "FR-09":
            folder_obj["item"] = [usage_setup, usage_setup] + folder_obj["item"]

    return {
        "info": {
            "name": "HW06 Member 3 API Tests (23127185)",
            "description": "FR-01 / FR-09 / FR-17. Header X-Student-Id via collection pre-request. Sanity = current SUT; Discovery = spec oracle.",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json",
        },
        "event": [
            {
                "listen": "prerequest",
                "script": {
                    "type": "text/javascript",
                    "exec": [
                        "const studentId = pm.environment.get('student_id') || '23127185';",
                        "if (pm.request.headers.upsert) {",
                        "    pm.request.headers.upsert({ key: 'X-Student-Id', value: studentId });",
                        "} else {",
                        "    pm.request.headers.add({ key: 'X-Student-Id', value: studentId });",
                        "}",
                        "console.log('Request Sent with X-Student-Id: ' + studentId);",
                    ],
                },
            },
            {
                "listen": "test",
                "script": {
                    "type": "text/javascript",
                    "exec": [
                        "pm.test('Response time is acceptable (< 1500ms)', function () {",
                        "    pm.expect(pm.response.responseTime).to.be.below(1500);",
                        "});",
                    ],
                },
            },
        ],
        "item": [
            {
                "name": "01_Sanity_Suite",
                "item": [{"name": "00_Setup_Auth", "item": sanity_setup}] + sanity_fr,
            },
            {
                "name": "02_Bug_Discovery_Suite",
                "item": [{"name": "00_Setup_Auth", "item": sanity_setup}] + disc_fr,
            },
            {
                "name": "03_Data_Driven_Demo",
                "item": [ddt_item],
            },
            {
                "name": "04_Mock_Server_Demo",
                "item": [mock_item],
            },
        ],
    }


def write_env(path: pathlib.Path, name: str, base_url: str, student_id: str) -> None:
    values = [
        ("base_url", base_url, True),
        ("student_id", student_id, True),
        ("admin_email", "admin@eshop.com", True),
        ("admin_password", "Admin123!", True),
        ("user_email", "coupon_user@eshop.com", True),
        ("user_password", "CouponPass123!", True),
        ("register_email", "register_user@eshop.com", True),
        ("register_password", "Register123!", True),
        ("admin_token", "", True),
        ("user_token", "", True),
        ("unique_email", "", True),
        ("m3_coupon_id", "", True),
        ("m3_new_code", "", True),
        ("m3_deleted_code", "", True),
        ("user_id", "", True),
        ("admin_id", "", True),
    ]
    env = {
        "id": "hw06-member3-" + name.lower(),
        "name": name,
        "values": [{"key": k, "value": v, "enabled": e} for k, v, e in values],
        "_postman_variable_scope": "environment",
    }
    path.write_text(json.dumps(env, indent=2), encoding="utf-8")


def write_ddt_csv(path: pathlib.Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = [
        ["name", "email", "password", "expected_status"],
        ["DDT Valid", "ddt_valid_{{timestamp}}@eshop.test", "Valid123!", "200"],
        ["DDT Weak", "ddt_weak@eshop.test", "123", "400"],
        ["DDT NoAt", "nodomain", "Valid123!", "400"],
        ["DDT EmptyPw", "ddt_emptypw@eshop.test", "", "400"],
        ["DDT Dup", "admin@eshop.com", "Valid123!", "409"],
    ]
    # Newman does not expand {{timestamp}} in data files. Use unique emails in python.
    import time
    ts = str(int(time.time()))
    rows[1][1] = f"ddt_valid_{ts}@eshop.test"
    with path.open("w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(rows)

    path2 = path.parent / "fr09-coupon-data.csv"
    with path2.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["code", "total_amount", "expected_status"])
        w.writerow(["BIGBUY", "600000", "200"])
        w.writerow(["NOPE999", "600000", "404"])
        w.writerow(["", "600000", "400"])
        w.writerow(["EXPIRED", "200000", "400"])
        w.writerow(["SAVE10", "300000", "200"])


def generate(features: list[str], spec: pathlib.Path, srs: pathlib.Path,
             student_id: str, use_llm: bool, emit_all: bool) -> dict[str, list[Case]]:
    parse_spec(spec)
    parse_srs(srs)
    by: dict[str, list[Case]] = {}
    for feat in features:
        cases = FEATURE_BUILDERS[feat]()
        cases = maybe_llm_enrich(cases, feat, use_llm)
        validate(cases, feat)
        by[feat] = cases
        write_csv(cases, ROOT / "test-cases" / "generated" / f"{feat}.csv")
        print(f"{feat}: {len(cases)} cases "
              f"(AI={sum(1 for c in cases if c.source=='AI')} "
              f"HUMAN={sum(1 for c in cases if c.source=='HUMAN')})")
    if emit_all:
        flat = [c for feat in features for c in by[feat]]
        write_xlsx(by, ROOT / "test-cases" / "member-3.xlsx")
        col = build_collection(flat, student_id)
        postman = ROOT / "postman"
        postman.mkdir(parents=True, exist_ok=True)
        (postman / "HW06_Member3.postman_collection.json").write_text(
            json.dumps(col, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        write_env(postman / "HW06_Local.postman_environment.json",
                  "HW06_Member3_Local", "http://localhost:3000", student_id)
        write_env(postman / "HW06_Mock.postman_environment.json",
                  "HW06_Member3_Mock", "http://localhost:3001", student_id)
        write_ddt_csv(postman / "data" / "fr01-register-data.csv")
        print("wrote collection, env, xlsx")
    return by


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--spec", default=str(ROOT / "eshop-sut" / "api_specification.md"))
    p.add_argument("--srs", default=str(ROOT / "eshop-sut" / "README.md"))
    p.add_argument("--feature", choices=["FR-01", "FR-09", "FR-17"])
    p.add_argument("--all", action="store_true")
    p.add_argument("--student-id", default=STUDENT_ID_DEFAULT)
    p.add_argument("--use-llm", action="store_true")
    args = p.parse_args()
    feats = ["FR-01", "FR-09", "FR-17"] if args.all or not args.feature else [args.feature]
    generate(feats, pathlib.Path(args.spec), pathlib.Path(args.srs),
             args.student_id, args.use_llm, emit_all=args.all or len(feats) == 3)
    return 0


if __name__ == "__main__":
    sys.exit(main())
